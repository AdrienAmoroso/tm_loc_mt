import os
import sys
import hashlib
from dataclasses import dataclass
from typing import List, Dict, Any
from pathlib import Path

# Add src directory to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from openai import OpenAI

import re
import json
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
import time
from datetime import datetime

from google import genai
from google.genai import types


# --- Config ---
SOURCE_LANG = "English"
TARGET_LANG = "Russian"   
TARGET_COL = "Russian"
DOC_SHEETS = [
    "Default",
    "Sandbox",
    "GEO",
    "TENNIS",
    "MANAGER",
    "EQUIP",
    "TRAINING",
    "MATCH",
    "STAFF",
    "INBOX",
    "PLAYER",
    "ACADEMY",
    "SCENARIOS",
    "TUTO",
    "FC",
    "MEDIA",
    "LD_INBOX",
    "LD_RADIO",
    "UI",
    "WORLD",
    "LD_BRIEFING",
    "LD_OBJ_CA",
    "LD_OBJ_PLY",
    "LD_RPT_TRAIN",
    "LD_TALK",
    "LD_ADVICES",
    "MP_UI",
    "MP_ATT",
    "MP_NEWS",
    "MP_SKILLS",
    "MP_TUTO"
    
]

SHEETS_TO_TRAD = [
    "TUTO"
]            

EXCEL_PATH = "localization.xlsx"

PLACEHOLDER_PATTERN_TAG = re.compile(r"<[^>]+>")
PLACEHOLDER_PATTERN_VAR = re.compile(r"{\[[^}]+\]}")

BATCH_COOLDOWN_SECONDS = 22.0
BATCH_SIZE = 20

USE_GEMINI = True

# --- OPENAI API ---
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
MODEL_NAME = "gpt-4o-mini"

# --- Gemini API ---
gemini_client = None
GEMINI_MODEL = "gemini-2.5-flash-lite"

if USE_GEMINI:
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY n'est pas défini dans les variables d'environnement")
    gemini_client = genai.Client(api_key=GEMINI_API_KEY)
else:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY n'est pas défini dans les variables d'environnement")




# --- Logs ---
LOGS_DIR = Path("logs")
LOGS_DIR.mkdir(exist_ok=True)

run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = LOGS_DIR / f"mt_run_{run_id}.log"
KEYS_LOG_FILE = LOGS_DIR / f"mt_keys_{run_id}.csv"

logging.basicConfig(
    level=logging.INFO,  # Back to INFO for cleaner output
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()  
    ]
)

logger = logging.getLogger(__name__)

# --- Segment Class---

@dataclass
class Segment:
    sheet: str
    row_idx: int        
    key: str
    source_text: str
    existing_target: str
    comment: str
    donottranslate: bool



def load_segments_from_sheet(path: str, sheet_name: str) -> List[Segment]:
    df = pd.read_excel(path, sheet_name=sheet_name)

    # Colonnes indispensables
    required_cols = ["Keys", SOURCE_LANG]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Missing column '{col}' in sheet '{sheet_name}'")

    has_target = TARGET_COL in df.columns
    if not has_target:
        logger.warning(f"[{sheet_name}] Target column '{TARGET_COL}' not found. Assuming empty targets for all rows.")
        
    # Colonnes optionnelles
    has_comment = "$comment" in df.columns
    has_donottranslate = "$donottranslate" in df.columns

    if not has_comment:
        logger.warning(f"[{sheet_name}] Column '$comment' not found. Comments will be empty.")
    if not has_donottranslate:
        logger.warning(f"[{sheet_name}] Column '$donottranslate' not found. Assuming donottranslate=False for all rows.")

    segments: List[Segment] = []

    for idx, row in df.iterrows():
        key = str(row["Keys"])

        source = str(row[SOURCE_LANG]) if not pd.isna(row[SOURCE_LANG]) else ""
        
        target = ""
        if has_target and not pd.isna(row[TARGET_COL]):
            target = str(row[TARGET_COL])

        comment = ""
        if has_comment and not pd.isna(row["$comment"]):
            comment = str(row["$comment"])

        donottranslate = False
        if has_donottranslate and not pd.isna(row["$donottranslate"]):
            donot = str(row["$donottranslate"]).strip().lower()
            donottranslate = (donot != "")

        seg = Segment(
            sheet=sheet_name,
            row_idx=idx + 2,  # header row = 1
            key=key,
            source_text=source,
            existing_target=target,
            comment=comment,
            donottranslate=donottranslate,
        )
        segments.append(seg)

    return segments





def should_translate(seg: Segment) -> bool:
    
    if not seg.source_text.strip():
        return False
    
    if seg.donottranslate:
        return False

    if seg.existing_target.strip():
        return False

    return True


def protect_placeholders(text: str) -> (str, Dict[str, str]):  # type: ignore
    placeholder_map: Dict[str, str] = {}

    # On va marquer d'abord les {[...]} puis les <...> sur le texte modifié
    def replace_with_tokens(pattern, prefix, input_text):
        matches = list(pattern.finditer(input_text))
        if not matches:
            return input_text

        out = []
        last = 0
        for i, m in enumerate(matches):
            out.append(input_text[last:m.start()])
            original = m.group(0)
            token = f"__{prefix}{i}__"
            placeholder_map[token] = original
            out.append(token)
            last = m.end()
        out.append(input_text[last:])
        return "".join(out)

    protected = replace_with_tokens(PLACEHOLDER_PATTERN_VAR, "VAR", text)
    protected = replace_with_tokens(PLACEHOLDER_PATTERN_TAG, "TAG", protected)
    return protected, placeholder_map

def tokens_in_order(text: str, tokens: List[str]) -> bool:
    pos = -1
    for t in tokens:
        new_pos = text.find(t, pos + 1)
        if new_pos == -1:
            return False
        pos = new_pos
    return True


def restore_placeholders(text: str, placeholder_map: Dict[str, str]) -> str:
    """Restore original placeholders from tokens with sorted iteration to ensure consistent order."""
    restored = text
    # Sort tokens to ensure deterministic replacement order (handles nested/overlapping patterns safely)
    for token in sorted(placeholder_map.keys()):
        original = placeholder_map[token]
        restored = restored.replace(token, original)
    return restored

def build_batches(segments: List[Segment], batch_size=BATCH_SIZE) -> List[List[Segment]]:
    to_translate = [s for s in segments if should_translate(s)]
    batches = []
    for i in range(0, len(to_translate), batch_size):
        batches.append(to_translate[i:i + batch_size])
    return batches

def build_system_prompt(source_lang: str, target_lang: str) -> str:
    return f"""
You are a professional localization engine for the video game "Tennis Manager",
a realistic tennis management game on PC. The player is the manager of a tennis academy
and manages up to 8 players while meeting sporting and financial objectives.

You translate in-game text from {source_lang} to {target_lang}.

GENERAL RULES
1. Technical tokens:
   - Never modify, translate or reorder technical tokens like __VAR0__, __VAR1__, __TAG0__, etc.
   - They represent placeholders or markup and must remain exactly as they are and in the same position.
   - If the source contains tokens like __TAG0__, __TAG1__, etc., the translation MUST contain exactly the same tokens, in the same order. Do not remove them, even if you change the wording.

2. Source format:
   - Texts come from an Excel localization sheet.
   - Each entry has:
     - a 'key' (identifier),
     - a 'sheet' name (e.g. "UI", "Staff", "Match", "Media", "Tuto", "Ld_talk", "Ld_advices", etc.),
     - an optional 'comment' that gives context.
   - Use the 'sheet' and 'comment' to adapt the tone.

3. Style and tone:
   - Overall tone: natural and fluent, consistent with a serious sports management game.
   - For sheets like "UI", "Default", "Geo", "Tennis", "Manager", "Equip":
     - These are UI labels or short texts → keep translations as short and clear as possible.
   - For "Tuto" and other long tutorial texts:
     - You may use slightly longer, didactic sentences, but still concise.
   - For "Match" and "Ld_radio":
     - These are live match comments → dynamic, natural speech, like a commentator.

4. Media and dialogues:
   - Sheet "Media": questions and answers in interviews and press conferences addressed to the manager.
     - Keep a tone that fits journalists (questions) and a manager answering them.
   - Sheet "Ld_talk": the manager talking to a player before or after a match (locker-room talk).
   - Sheet "Ld_advices": short tips or advice given between sets.
   - Make sure the tone matches the situation: professional, motivational, not slangy.

5. Gender handling (_M / _F keys):
   - If the key ends with "_M", the text refers to a male character (player, manager or staff).
   - If the key ends with "_F", the text refers to a female character.
   - Only adapt gender where the text actually refers to the person (player/manager/staff), not for generic parts.
   - If there is no _M/_F suffix and gender is ambiguous, prefer a neutral form when possible or the generic masculine, depending on what is most natural in {target_lang}.

6. General constraints:
   - Do not add any new meaning, do not omit important information.
   - Keep roughly the same level of formality as the English source (no slang, no overly formal expressions).
   - Avoid making the text significantly longer than the source, especially for UI-related sheets.

LAYOUT / LINE BREAKS
- Preserve line breaks exactly. If the source contains multiple lines separated by "\n", keep the same number of lines in the same order.
- If a line starts with a bullet/marker (e.g. "-", "•", "1.", "2."), keep the same marker at the start of the translated line.
- Do not merge lines. Do not split a line into multiple lines.

RESPONSE FORMAT
- Do NOT add explanations or comments.
- Answer strictly as valid JSON with this structure:
{{
  "translations": [
    {{ "key": "KEY_FROM_INPUT", "text": "Translated text here" }},
    {{ "key": "ANOTHER_KEY", "text": "Another translation" }}
  ]
}}
- Return exactly as many translations as segments were provided in the INPUT.
- Preserve all keys exactly as provided.
"""

def build_user_content(batch: List[Segment]) -> str:
    payload = {
        "segments": []
    }
    for seg in batch:
        protected, _ = protect_placeholders(seg.source_text)
        payload["segments"].append({
            "key": seg.key,
            "sheet": seg.sheet,
            "source": protected,
            "comment": seg.comment,
        })
    return json.dumps(payload, ensure_ascii=False)


def call_openai_for_batch(
    batch: List[Segment],
    max_retries: int = 5,
    default_wait: float = 25.0
) -> Dict[str, str]:
    """
    Appelle l'API OpenAI pour un batch de segments.
    Gère les erreurs réseau / rate limit avec un retry simple.
    Retourne un dict key -> texte_traduit (placeholders encore protégés).
    """
    system_prompt = build_system_prompt(SOURCE_LANG, TARGET_LANG)
    user_content = build_user_content(batch)

    last_exception = None

    for attempt in range(1, max_retries + 1):
        try:
            logger.info(
                f"Calling OpenAI for batch of {len(batch)} segments "
                f"(attempt {attempt}/{max_retries})"
            )

            response = openai_client.responses.create(
                model=MODEL_NAME,
                input=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content}
                ]
            )

            raw_text = response.output[0].content[0].text
            result = parse_translations_json(raw_text)
            return result

        except json.JSONDecodeError as e:
            logger.error(f"[API] JSON parsing error, OpenAI response: {e}")
            last_exception = e

        except Exception as e:
            msg = str(e)
            last_exception = e

            # Rate limit (429)
            if "rate_limit_exceeded" in msg or "Rate limit reached" in msg:
                import re
                wait_seconds = default_wait
                m = re.search(r"try again in (\d+)s", msg)
                if m:
                    wait_seconds = max(default_wait, float(m.group(1)) + 5.0)

                logger.warning(
                    f"[API] Rate limit atteint (attempt {attempt}/{max_retries}). "
                    f"Waiting {wait_seconds:.1f}s before retry..."
                )

                if attempt == max_retries:
                    break  

                time.sleep(wait_seconds)
                continue  

            wait_seconds = 5.0 * attempt
            logger.error(
                f"[API] Error calling OpenAI (attempt {attempt}/{max_retries}): {e}. "
                f"Retrying in {wait_seconds:.1f}s..."
            )

            if attempt == max_retries:
                break

            time.sleep(wait_seconds)

    
    logger.error("[API] All OpenAI calls failed attempt")
    if last_exception:
        raise last_exception
    else:
        raise RuntimeError("Unknown error calling OpenAI API.")

def parse_translations_json(raw_text: str) -> Dict[str, str]:
    """Parse JSON response from API and extract translations."""
    data = json.loads(raw_text)
    
    if "translations" not in data:
        logger.error(f"[Parse] 'translations' key missing from response. Keys found: {list(data.keys())}")
        raise ValueError("Invalid JSON format: missing 'translations' list")
    
    if not isinstance(data["translations"], list):
        logger.error(f"[Parse] 'translations' is not a list, got: {type(data['translations'])}")
        raise ValueError("Invalid JSON format: 'translations' is not a list")

    result: Dict[str, str] = {}
    for item in data["translations"]:
        if not isinstance(item, dict):
            continue
        k = item.get("key")
        txt = item.get("text", "")
        if k:
            result[str(k)] = str(txt)
    
    if not result:
        logger.warning(f"[Parse] Parsed translations list but no valid key/text pairs found. "
                      f"Translations list length: {len(data.get('translations', []))}")
    
    return result


def call_gemini_for_batch(
    batch: List[Segment],
    max_retries: int = 5,
    sleep_base: float = 5.0,
    default_wait: float = 25.0
) -> Dict[str, str]:
    """
    Appelle Gemini pour un batch de segments
    Retourne un dict key -> texte_traduit (placeholders encore protégés).

    batch : liste de Segment (même structure que pour OpenAI).
    """

    system_prompt = build_system_prompt(SOURCE_LANG, TARGET_LANG)   
    user_content = build_user_content(batch)
    
    logger.debug(f"[Gemini] User content first 300 chars: {user_content[:300]}")           
    logger.debug(f"[Gemini] Full prompt length: {len(system_prompt) + len(user_content)}")
    
    full_prompt = system_prompt + "\n\n" + user_content

    gen_config = types.GenerateContentConfig(
        temperature=0.0,
        candidate_count=1,
        response_mime_type="application/json",  
    )
    
    if gemini_client is None:
        raise RuntimeError("Gemini client not initialized (USE_GEMINI is False or key missing).")

    for attempt in range(1, max_retries + 1):
        try:
            logger.info(
                f"[Gemini] Calling {GEMINI_MODEL} for batch of "
                f"{len(batch)} segments (attempt {attempt}/{max_retries})"
            )
            
            # Debug: log the segments being sent
            logger.debug(f"[Gemini] Batch segments: {[(s.key, s.source_text[:30]) for s in batch]}")

            response = gemini_client.models.generate_content(
                model=GEMINI_MODEL,
                contents=full_prompt,
                config=gen_config,
            )

            raw_text = response.text
            logger.debug(f"[Gemini] Response object encoding: {getattr(response, 'encoding', 'N/A')}")
            logger.debug(f"[Gemini] Raw API response (first 1000 chars): {raw_text[:1000]}")
            result = parse_translations_json(raw_text)
            
            if not result:
                logger.warning(f"[Gemini] API returned empty translations dict for batch of {len(batch)} segments. "
                              f"Raw response: {raw_text[:500]}")
            
            return result

        except json.JSONDecodeError as e:
            logger.error(f"[Gemini] JSON parsing error, response: {e}")
            logger.error(f"[Gemini] Raw response was: {getattr(response, 'text', 'NO_TEXT')[:500]}")
            last_exception = e

        except Exception as e:
            msg = str(e)
            last_exception = e

            if "429" in msg or "RESOURCE_EXHAUSTED" in msg or "rate limit" in msg.lower():
                wait_seconds = default_wait
                import re
                m = re.search(r"(\d+)\s*s", msg)
                if m:
                    wait_seconds = max(default_wait, float(m.group(1)) + 5.0)

                logger.warning(
                    f"[Gemini] Rate limit/quota atteint (attempt {attempt}/{max_retries}), "
                    f"waiting {wait_seconds:.1f}s before retry..."
                )
                time.sleep(wait_seconds)
                continue

            logger.error(f"[Gemini] Erreur lors de l'appel Gemini (attempt {attempt}/{max_retries}): {e}")
            sleep_time = sleep_base * attempt
            logger.info(f"[Gemini] Retrying in {sleep_time:.1f}s...")
            time.sleep(sleep_time)

    logger.error("[Gemini] Tous les essais d'appel API ont échoué pour ce batch.")
    if last_exception:
        raise last_exception
    else:
        raise RuntimeError("Unknown error calling Gemini API.")


def call_model_for_batch(batch: List[Segment]) -> Dict[str, str]:
    if USE_GEMINI:
        return call_gemini_for_batch(batch)
    return call_openai_for_batch(batch)

def append_key_log(path: Path, sheet: str, key: str, row_idx: int, target_lang: str, status: str):
    new_file = (not path.exists()) or (path.stat().st_size == 0)
    with path.open("a", encoding="utf-8") as f:
        if new_file:
            f.write("sheet,key,row_idx,target_lang,status\n")
        f.write(f"{sheet},{key},{row_idx},{target_lang},{status}\n")


def apply_translations_to_segments(
    segments: List[Segment],
    translations: Dict[str, str],
    keys_log_path: Path
) -> Dict[int, str]:
    """
    Returns a mapping row_idx -> final translated text to be written into Excel.
    Also logs each processed key into a CSV file.
    """
    row_to_text: Dict[int, str] = {}

    for seg in segments:
        # Handle donottranslate segments: copy source text directly
        if seg.donottranslate:
            source = seg.source_text
            if source and source.strip():
                row_to_text[seg.row_idx] = source
                logger.info(
                    f"[MT] Copying source '{SOURCE_LANG}' for donottranslate key={seg.key} row={seg.row_idx}"
                )
                append_key_log(keys_log_path, seg.sheet, seg.key, seg.row_idx, TARGET_LANG, "COPIED_SOURCE")
            continue

        # Skip segments that should not be translated (empty source, already has target, etc.)
        if not should_translate(seg):
            continue

        # If the model didn't return a translation for this key, log it and skip
        translated_protected = translations.get(seg.key, "")
        if not translated_protected or not translated_protected.strip():
            logger.warning(f"[MT] No translation returned for key={seg.key} (sheet={seg.sheet})")
            append_key_log(keys_log_path, seg.sheet, seg.key, seg.row_idx, TARGET_LANG, "NO_TRANSLATION")
            continue

        # Protect placeholders in the source, so we can validate and restore them after translation
        protected_source, placeholder_map = protect_placeholders(seg.source_text)

        # Tokens in the real order of appearance in the protected source
        tokens_in_source = re.findall(r"__(?:VAR|TAG)\d+__", protected_source)

        # Validate: all tokens must still be present in the translated text
        missing_tokens = [tok for tok in tokens_in_source if tok not in translated_protected]
        if missing_tokens:
            logger.warning(
                f"[MT] Missing tokens {missing_tokens} for key={seg.key} (sheet={seg.sheet})"
            )
            append_key_log(keys_log_path, seg.sheet, seg.key, seg.row_idx, TARGET_LANG, "MISSING_TOKENS")
            # Safer: do not write a broken string into Excel
            continue

        # Validate: tokens should keep the same order
        if tokens_in_source and not tokens_in_order(translated_protected, tokens_in_source):
            logger.warning(
                f"[MT] Tokens out of order for key={seg.key} (sheet={seg.sheet}). "
                f"Skipping to avoid corruption."
            )
            append_key_log(keys_log_path, seg.sheet, seg.key, seg.row_idx, TARGET_LANG, "TOKENS_OUT_OF_ORDER")
            continue

        # Restore original placeholders / tags
        final_text = restore_placeholders(translated_protected, placeholder_map)

        # Store result for Excel write
        row_to_text[seg.row_idx] = final_text

        logger.info(f"[MT] Translated key={seg.key} sheet={seg.sheet} row={seg.row_idx} status=OK")
        append_key_log(keys_log_path, seg.sheet, seg.key, seg.row_idx, TARGET_LANG, "OK")

    return row_to_text




def find_column_index(ws, target_col_name: str) -> int:
    header_row = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == target_col_name:
            return col
    return -1

def ensure_column_exists(ws, target_col_name: str) -> int:
    col_index = find_column_index(ws, target_col_name)
    if col_index != -1:
        return col_index

    # crée la colonne à droite
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=target_col_name)
    logger.warning(f"[Excel] Column '{target_col_name}' not found. Created at column {new_col} in sheet '{ws.title}'.")
    return new_col

def write_updates_in_ws(ws, row_to_text: Dict[int, str], target_col_name: str):
    col_index = ensure_column_exists(ws, target_col_name)
    for row_idx, text in row_to_text.items():
        ws.cell(row=row_idx, column=col_index, value=text)


def verify_and_fill_gaps(excel_path: str, sheet_name: str, source_col: str, target_col: str, keys_log_path: Path) -> int:
    """
    Verify that all translatable rows have translations.
    Fill gaps for rows with empty target but non-empty source.
    
    Returns: number of gaps filled
    """
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    
    # Check for required columns
    if "Keys" not in df.columns or source_col not in df.columns:
        logger.warning(f"[Verify] Missing Keys or {source_col} column in sheet {sheet_name}")
        return 0
    
    has_target = target_col in df.columns
    if not has_target:
        logger.warning(f"[Verify] Target column '{target_col}' not found. Nothing to verify.")
        return 0
    
    has_comment = "$comment" in df.columns
    has_donottranslate = "$donottranslate" in df.columns
    
    # Find gaps: rows where source is not empty but target is empty
    gaps: List[Segment] = []
    
    for idx, row in df.iterrows():
        key = str(row["Keys"])
        source = str(row[source_col]) if not pd.isna(row[source_col]) else ""
        target = str(row[target_col]) if not pd.isna(row[target_col]) else ""
        
        # Skip if source is empty
        if not source or not source.strip():
            continue
        
        # Skip if target is already filled
        if target and target.strip():
            continue
        
        # Skip if marked as donottranslate
        donottranslate = False
        if has_donottranslate and not pd.isna(row["$donottranslate"]):
            donot = str(row["$donottranslate"]).strip().lower()
            donottranslate = (donot != "")
        
        if donottranslate:
            continue
        
        # This is a gap - source exists but target is empty
        comment = ""
        if has_comment and not pd.isna(row["$comment"]):
            comment = str(row["$comment"])
        
        seg = Segment(
            sheet=sheet_name,
            row_idx=idx + 2,  # header row = 1
            key=key,
            source_text=source,
            existing_target="",
            comment=comment,
            donottranslate=False,
        )
        gaps.append(seg)
    
    if not gaps:
        logger.info(f"[Verify] No gaps found in sheet '{sheet_name}'")
        return 0
    
    logger.info(f"[Verify] Found {len(gaps)} gaps in sheet '{sheet_name}'. Attempting to fill...")
    
    # Build and process gap batches
    gap_batches = build_batches(gaps, batch_size=BATCH_SIZE)
    gap_count = 0
    
    for i, batch in enumerate(gap_batches, start=1):
        logger.info(f"[Verify] Processing gap batch {i}/{len(gap_batches)} with {len(batch)} segments")
        
        try:
            translations = call_model_for_batch(batch)
        except Exception as e:
            logger.error(f"[Verify] Failed to get translations for gap batch {i}: {e}")
            continue
        
        # Apply gap translations
        gap_row_updates = apply_translations_to_segments(batch, translations, keys_log_path)
        
        if gap_row_updates:
            # Write gap translations to Excel
            ws = load_workbook(excel_path)[sheet_name]
            write_updates_in_ws(ws, gap_row_updates, target_col)
            gap_count += len(gap_row_updates)
            
            logger.info(f"[Verify] Filled {len(gap_row_updates)} gaps in gap batch {i}")
        
        # Cooldown between batches
        if i < len(gap_batches):
            logger.info(f"[Verify] Cooldown {BATCH_COOLDOWN_SECONDS:.1f}s before next gap batch...")
            time.sleep(BATCH_COOLDOWN_SECONDS)
    
    logger.info(f"[Verify] Total gaps filled: {gap_count}")
    return gap_count

    

def main():
    logger.info(f"=== MT run started for target_lang={TARGET_LANG}, target_col={TARGET_COL} ===")
    wb = load_workbook(EXCEL_PATH)

    existing_sheets = set(wb.sheetnames)
    
    for sheet_name in SHEETS_TO_TRAD:
        if sheet_name not in existing_sheets:
            logger.error(f"[Main] Sheet '{sheet_name}' not found in workbook. Skipping.")
            continue
        logger.info(f"--- Processing sheet '{sheet_name}' ---")

        segments = load_segments_from_sheet(EXCEL_PATH, sheet_name)
        batches = build_batches(segments, batch_size=BATCH_SIZE)

        logger.info(f"[{sheet_name}] Total segments in sheet: {len(segments)}")
        logger.info(f"[{sheet_name}] Segments to translate (should_translate): "
                    f"{sum(1 for s in segments if should_translate(s))}")
        logger.info(f"[{sheet_name}] Number of batches: {len(batches)}")

        all_row_updates: Dict[int, str] = {}

        for i, batch in enumerate(batches, start=1):
            logger.info(f"[{sheet_name}] Processing batch {i}/{len(batches)} with {len(batch)} segments")

            try:
                translations = call_model_for_batch(batch)
            except Exception as e:
                logger.error(f"[{sheet_name}][BATCH {i}] Failed to get translations from API: {e}")
                break

            row_to_text = apply_translations_to_segments(batch, translations, KEYS_LOG_FILE)
            all_row_updates.update(row_to_text)
            
            logger.info(
                f"[{sheet_name}] Cooldown {BATCH_COOLDOWN_SECONDS:.1f}s before next batch "
                f"to respect rate limits..."
            )
            time.sleep(BATCH_COOLDOWN_SECONDS)

        logger.info(f"[{sheet_name}] Total rows updated: {len(all_row_updates)}")

        if all_row_updates:
            ws = wb[sheet_name]
            write_updates_in_ws(ws, all_row_updates, TARGET_COL)
        else:
            logger.info(f"[{sheet_name}] No updates to write.")

    wb.save(EXCEL_PATH)
    logger.info(f"[Excel] Saved all updates into {EXCEL_PATH}")

    # Verification phase: fill any remaining gaps
    logger.info("=== Starting verification phase ===")
    total_gaps_filled = 0
    for sheet_name in SHEETS_TO_TRAD:
        if sheet_name not in existing_sheets:
            continue
        gaps_filled = verify_and_fill_gaps(EXCEL_PATH, sheet_name, SOURCE_LANG, TARGET_COL, KEYS_LOG_FILE)
        total_gaps_filled += gaps_filled
    
    if total_gaps_filled > 0:
        # Save again if gaps were filled
        wb = load_workbook(EXCEL_PATH)
        wb.save(EXCEL_PATH)
        logger.info(f"[Excel] Saved gap-filled updates into {EXCEL_PATH}")
    
    logger.info(f"=== MT run finished (total gaps filled: {total_gaps_filled}) ===")





if __name__ == "__main__":
    main()
