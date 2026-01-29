"""Excel file operations for reading and writing translations."""

import logging

import pandas as pd
from openpyxl import load_workbook

from config import Config
from models import Segment

logger = logging.getLogger(__name__)


class ExcelService:
    """Handles reading and writing of Excel localization files."""

    def __init__(self, config: Config):
        self.config = config

    def load_segments_from_sheet(self, sheet_name: str) -> list[Segment]:
        """Load all segments from an Excel sheet."""
        df = pd.read_excel(str(self.config.excel.excel_path), sheet_name=sheet_name)

        # Validate required columns
        required_cols = [self.config.excel.keys_column, self.config.translation.source_lang]
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Missing column '{col}' in sheet '{sheet_name}'")

        # Check for optional columns
        has_target = self.config.translation.target_col in df.columns
        has_comment = self.config.excel.comment_column in df.columns
        has_donottranslate = self.config.excel.donottranslate_column in df.columns

        if not has_target:
            logger.warning(
                f"[{sheet_name}] Target column '{self.config.translation.target_col}' not found"
            )
        if not has_comment:
            logger.warning(f"[{sheet_name}] Comment column not found")
        if not has_donottranslate:
            logger.warning(f"[{sheet_name}] DoNotTranslate column not found")

        segments = []

        for idx, row in df.iterrows():
            key = str(row[self.config.excel.keys_column])
            source = (
                str(row[self.config.translation.source_lang])
                if pd.notna(row[self.config.translation.source_lang])
                else ""
            )
            target = (
                str(row[self.config.translation.target_col])
                if (has_target and pd.notna(row[self.config.translation.target_col]))
                else ""
            )
            comment = (
                str(row[self.config.excel.comment_column])
                if (has_comment and pd.notna(row[self.config.excel.comment_column]))
                else ""
            )

            donottranslate = False
            if has_donottranslate and pd.notna(row[self.config.excel.donottranslate_column]):
                donot = str(row[self.config.excel.donottranslate_column]).strip().lower()
                donottranslate = donot != ""

            seg = Segment(
                sheet=sheet_name,
                row_idx=idx + 2,  # header row = 1
                key=key,
                source_text=source,
                existing_target=target,
                comment=comment,
                donottranslate=donottranslate,
            )
            segments.append(seg)

        logger.info(f"[{sheet_name}] Loaded {len(segments)} segments")
        return segments

    def find_column_index(self, ws, target_col_name: str) -> int:
        """Find column index by header name."""
        header_row = 1
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col).value == target_col_name:
                return col
        return -1

    def ensure_column_exists(self, ws, target_col_name: str) -> int:
        """Ensure target column exists, create if needed."""
        col_index = self.find_column_index(ws, target_col_name)
        if col_index != -1:
            return col_index

        # Create new column
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value=target_col_name)
        logger.warning(f"[Excel] Created column '{target_col_name}' at column {new_col}")
        return new_col

    def write_translations(self, sheet_name: str, translations: dict[int, str]) -> None:
        """Write translations to Excel file."""
        if not translations:
            logger.info(f"[{sheet_name}] No translations to write")
            return

        wb = load_workbook(str(self.config.excel.excel_path))
        ws = wb[sheet_name]

        col_index = self.ensure_column_exists(ws, self.config.translation.target_col)

        for row_idx, text in translations.items():
            ws.cell(row=row_idx, column=col_index, value=text)

        wb.save(str(self.config.excel.excel_path))
        logger.info(f"[{sheet_name}] Wrote {len(translations)} translations to Excel")

    def get_existing_sheets(self) -> set:
        """Get list of all sheet names in the workbook."""
        wb = load_workbook(str(self.config.excel.excel_path))
        return set(wb.sheetnames)
